import openpyxl
filePath ="D:\pythonProject\oilprice3\OilPriceForecastingSoftware\python310\data.xlsx"
wb = openpyxl.load_workbook(filePath.replace("/", "//"))  # 读取文件路径
# 打开指定的工作簿中的指定工作表：
line = 2
ws = wb["Sheet1"]
ws = wb.active  # 打开激活的工作表
ws = list(ws.values)  # 转为列表
dates = list(ws[0])[5:]
base_information = list(ws[int(line) - 1])[:5]
print(base_information)
prices = list(ws[int(line) - 1])[5:]
print(prices)
prewaterlevel1 = list(ws[int(line)])[5:]
print(prewaterlevel1)
prewaterlevel2 = list(ws[int(line) + 1])[5:]
print(prewaterlevel2)