# 先打开源文件
import sys
import zipfile
import LSTMARIMA
import KalmanBP
import ARIMARF
import EMDARIMA
from PyQt5.QtWidgets import QApplication, QMainWindow, QVBoxLayout, QPushButton, QWidget, QSizePolicy
from PyQt5 import uic
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QPixmap, QPainter
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QTextEdit, QFileDialog, QComboBox, QHBoxLayout, QLabel, \
    QLineEdit, QVBoxLayout, QMainWindow

import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
import os


# 创建一个子线程,用来加载数据
class MyThread_getdata(QThread):
    # 导入一个信号，用来接受读取的
    getdata_signal = pyqtSignal(str, str)

    def __init__(self):
        super().__init__()

    def get_picture(self, dates, prices):
        ##设定一些基本参数
        os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
        plt.rcParams['axes.unicode_minus'] = False

        plt.rcParams['font.sans-serif'] = ['SimHei']
        # 创建DataFrame
        df = pd.DataFrame({'Date': eval(dates), 'Price': eval(prices)})
        df['Date'] = pd.to_datetime(df['Date'])
        # print(df['Date'])
        df.set_index('Date', inplace=True)

        # 绘制时序图
        plt.figure(figsize=(12, 6))
        plt.plot(df.index, df['Price'], marker='o')
        plt.title('长江成品油运价时序图')
        plt.xlabel('日期')
        plt.ylabel('价格')
        plt.grid(True)
        plt.savefig("output\\data_initial.png")

    def run(self):
        while True:
            pass


# 创建第二个子线程，用来训练ARIMA模型
class MyThread_ARIMA(QThread):
    # 导入一个信号，用来接受读取的
    ARIMA_signal = pyqtSignal(str, str, str, str, str)

    def __init__(self):
        super().__init__()

    def get_model(self, dates, prices, date_time, waterlevel1, waterlevel2):
        LSTMARIMA.main(eval(dates), eval(prices), date_time, waterlevel1, waterlevel2)

    def run(self):
        while True:
            pass


# 创建第三个子线程，用来使用ARIMA预测
class MyThread_ARIMA_pre(QThread):
    # 导入一个信号，用来接受读取的
    ARIMA_pre_signal = pyqtSignal(str, str)

    def __init__(self):
        super().__init__()

    def model_pre(self, dates, prices):
        # LSTMARIMA.pre(eval(dates), eval(prices), date_time)
        pass

    def run(self):
        while True:
            pass


# 创建第四个子线程，用来训练KalmanBP模型
class MyThread_KalmanBP(QThread):
    # 导入一个信号，用来接受读取的
    KalmanBP_signal = pyqtSignal(str)

    def __init__(self):
        super().__init__()

    def get_model(self, out):
        KalmanBP.main(out)

    def run(self):
        while True:
            pass


# 创建第五个子线程，用来使用KalmanBP预测
class MyThread_KalmanBP_pre(QThread):
    # 导入一个信号，用来接受读取的
    KalmanBP_pre_signal = pyqtSignal(str, str)

    def __init__(self):
        super().__init__()

    def model_pre(self, dates, prices):
        pass

    def run(self):
        while True:
            pass


# 创建第六个子线程，用来训练KalmanBP模型
class MyThread_ARIMABF(QThread):
    # 导入一个信号，用来接受读取的
    ARIMABF_signal = pyqtSignal(str, str, str, str, str)

    def __init__(self):
        super().__init__()

    def get_model(self, dates, prices, date_time, waterlevel1, waterlevel2):
        ARIMARF.main(eval(dates), eval(prices), date_time, waterlevel1, waterlevel2)

    def run(self):
        while True:
            pass


# 创建第七个子线程，用来使用KalmanBP预测
class MyThread_ARIMABF_pre(QThread):
    # 导入一个信号，用来接受读取的
    ARIMABF_pre_signal = pyqtSignal(str, str)

    def __init__(self):
        super().__init__()

    def model_pre(self, dates, prices):
        pass

    def run(self):
        while True:
            pass





# 创建第八个子线程，用来训练KalmanBP模型
class MyThread_EMDARIMA(QThread):
    # 导入一个信号，用来接受读取的
    EMDARIMA_signal = pyqtSignal(str, str, str, str, str)

    def __init__(self):
        super().__init__()

    def get_model(self, dates, prices, date_time, waterlevel1, waterlevel2):
        EMDARIMA.main(eval(dates), eval(prices), date_time, waterlevel1, waterlevel2)

    def run(self):
        while True:
            pass


# 创建第九个子线程，用来使用KalmanBP预测
class MyThread_EMDARIMA_pre(QThread):
    # 导入一个信号，用来接受读取的
    EMDARIMA_pre_signal = pyqtSignal(str, str)

    def __init__(self):
        super().__init__()

    def model_pre(self, dates, prices,):
        # print(date_time)
        pass

    def run(self):
        while True:
            pass




class Mywindow(QWidget):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("成品油运价预测")
        # self.resize(600, 400)
        self.setMinimumSize(600, 400)
        self.msg_history = list()
        # self.setObjectName('win')  # 设置窗口名，相当于CSS中的ID
        # self.setStyleSheet('#win{border-image:url(background.png);}')  # 设置图片的相对路径
        self.init_ui()


    #设置窗口背景
    def paintEvent(self, event):
        painter = QPainter(self)
        # painter.drawPixmap(self.rect(), QPixmap('background.png'))  # 设置窗口背景图片，平铺到整个窗口，随着窗口的变化而变化
        painter.drawPixmap(self.rect(), QPixmap('background.jpg'))  # 设置窗口背景图片，平铺到整个窗口，随着窗口的变化而变化



    def init_ui(self):
        # 创建一个垂直布局器
        self.container = QHBoxLayout()

        # 创建按钮
        self.but_input = QPushButton("打开文件", self)
        self.but_input.setMaximumWidth(100)
        self.but_input.setStyleSheet(
            "QPushButton { background-color: #4CAF50; color: white; border-style: outset; border-width: 2px; border-radius: 100px; border-color: #2E8B57; }"
            "QPushButton:hover { background-color: #45a049; }"
            "QPushButton:pressed { background-color: #3e8e41; }")
        # self.but_input.setStyleSheet("QPushButton { border-radius: 15px; padding: 10px; background-color: #6495ED; color: white;}")
        # self.but_input.setStyleSheet("QPushButton { border-radius: 15px; padding: 10px; background-color: #4CAF50; color: white;}")
        # self.but_input.setFixedSize(150, 50)

        self.but_pred = QPushButton("开始训练", self)
        self.but_pred.setMaximumWidth(100)
        self.but_pred.setStyleSheet(
            "QPushButton { background-color: #4CAF50; color: white; border-style: outset; border-width: 2px; border-radius: 100px; border-color: #2E8B57; }"
            "QPushButton:hover { background-color: #45a049; }"
            "QPushButton:pressed { background-color: #3e8e41; }")
        # self.but_pred.setStyleSheet("QPushButton { border-radius: 15px; padding: 10px; background-color: #6495ED; color: white;}")
        # self.but_pred.setFixedSize(150, 50)

        self.but_load = QPushButton("加载数据", self)
        self.but_load.setMaximumWidth(100)
        self.but_load.setStyleSheet(
            "QPushButton { background-color: #4CAF50; color: white; border-style: outset; border-width: 2px; border-radius: 100px; border-color: #2E8B57; }"
            "QPushButton:hover { background-color: #45a049; }"
            "QPushButton:pressed { background-color: #3e8e41; }")
        # self.but_load.setStyleSheet("QPushButton { border-radius: 15px; padding: 10px; background-color: #6495ED; color: white;}")
        # self.but_load.setFixedSize(150, 50)

        self.but_img = QPushButton("显示训练结果", self)
        self.but_img.setMaximumWidth(100)
        self.but_img.setStyleSheet(
            "QPushButton { background-color: #4CAF50; color: white; border-style: outset; border-width: 2px; border-radius: 100px; border-color: #2E8B57; }"
            "QPushButton:hover { background-color: #45a049; }"
            "QPushButton:pressed { background-color: #3e8e41; }")
        # self.but_img.setStyleSheet("QPushButton { border-radius: 15px; padding: 10px; background-color: #6495ED; color: white;}")
        # self.but_img.setFixedSize(150, 50)

        self.but_out = QPushButton("导出数据", self)
        self.but_out.setStyleSheet(
            "QPushButton { background-color: #4CAF50; color: white; border-style: outset; border-width: 2px; border-radius: 100px; border-color: #2E8B57; }"
            "QPushButton:hover { background-color: #45a049; }"
            "QPushButton:pressed { background-color: #3e8e41; }")
        # self.but_out.setMinimumWidth(200)
        # self.but_out.setFixedSize(200, 50)

        self.but_nextyear_pre = QPushButton("开始预测", self)
        self.but_nextyear_pre.setMaximumWidth(100)
        self.but_nextyear_pre.setStyleSheet(
            "QPushButton { background-color: #4CAF50; color: white; border-style: outset; border-width: 2px; border-radius: 100px; border-color: #2E8B57; }"
            "QPushButton:hover { background-color: #45a049; }"
            "QPushButton:pressed { background-color: #3e8e41; }")
        # self.but_nextyear_pre.setStyleSheet("QPushButton { border-radius: 15px; padding: 10px; background-color: #6495ED; color: white;}")
        # self.but_nextyear_pre.setFixedSize(150, 50)

        self.but_nextyear_show = QPushButton("显示预测结果", self)
        self.but_nextyear_show.setMaximumWidth(100)
        self.but_nextyear_show.setStyleSheet(
            "QPushButton { background-color: #4CAF50; color: white; border-style: outset; border-width: 2px; border-radius: 100px; border-color: #2E8B57; }"
            "QPushButton:hover { background-color: #45a049; }"
            "QPushButton:pressed { background-color: #3e8e41; }")
        # self.but_nextyear_show.setStyleSheet("QPushButton { border-radius: 15px; padding: 10px; background-color: #6495ED; color: white;}")
        # self.but_nextyear_show.setFixedSize(150, 50)

        # 创建标签框
        self.image_label = QLabel(self)
        self.image_label.setText("图形绘制：")
        self.image_label.setStyleSheet(
            "QLabel { border-style: solid; border-width: 2px; border-color: #D3D3D3; color: #D3D3D3; }")
        # self.image_label.setFixedSize(, 300)

        self.label1 = QLabel(self)
        self.label1.setMaximumWidth(25)
        # self.label1.setFixedSize(50, 20)
        self.label1.setText("数据：")

        self.label2 = QLabel(self)
        self.label2.setMaximumWidth(10)
        # self.label2.setFixedSize(40, 20)
        self.label2.setText("行")

        self.label3 = QLabel(self)
        self.label3.setMaximumWidth(90)

        # self.label2.setFixedSize(40, 20)
        self.label3.setText("预测起始时间：")

        self.label4 = QLabel(self)
        self.label4.setMaximumWidth(5)
        # self.label2.setFixedSize(40, 20)
        self.label4.setText("—")  # 用来显示年月之间

        self.label5 = QLabel(self)
        self.label5.setMaximumWidth(5)
        # self.label2.setFixedSize(40, 20)
        self.label5.setText("—")  # 用来显示月日之间

        # 创建文本栏
        self.dateselect_lineEdit = QLineEdit()
        self.dateselect_lineEdit.setMaximumWidth(40)
        self.dateselect_lineEdit.setText("2")
        self.dateselect_lineEdit.setStyleSheet(
            "QLineEdit { background-color: white; color: #333; border-style: solid; border-width: 2px; border-color: #ccc; border-radius: 5px; padding: 3px; }")

        self.year_lineEdit = QLineEdit()
        self.year_lineEdit.setMaximumWidth(50)
        self.year_lineEdit.setText("2024")
        self.year_lineEdit.setStyleSheet(
            "QLineEdit { background-color: white; color: #333; border-style: solid; border-width: 2px; border-color: #ccc; border-radius: 5px; padding: 3px; }")

        self.mouth_lineEdit = QLineEdit()
        self.mouth_lineEdit.setMaximumWidth(25)
        self.mouth_lineEdit.setText("01")
        self.mouth_lineEdit.setStyleSheet(
            "QLineEdit { background-color: white; color: #333; border-style: solid; border-width: 2px; border-color: #ccc; border-radius: 5px; padding: 3px; }")

        self.day_lineEdit = QLineEdit()
        self.day_lineEdit.setMaximumWidth(30)
        self.day_lineEdit.setText("01")
        self.day_lineEdit.setStyleSheet(
            "QLineEdit { background-color: white; color: #333; border-style: solid; border-width: 2px; border-color: #ccc; border-radius: 5px; padding: 3px; }")

        # self.dateselect_lineEdit.setFixedSize(60, 40)
        self.textedit = QTextEdit()
        self.textedit.setMaximumHeight(50)
        self.textedit.setPlaceholderText("状态显示栏：")
        self.textedit.setStyleSheet(
            "QTextEdit { background-color: #f0f0f0; color: #333; border-style: solid; border-width: 2px; border-color: #ccc; border-radius: 5px; }")
        self.out_textedit = QTextEdit()
        self.out_textedit.setStyleSheet("font-size: 6pt;")
        self.out_textedit.setFixedWidth(150)
        self.out_textedit.setPlaceholderText("结果输出栏：")
        self.out_textedit.setStyleSheet(
            "QTextEdit { background-color: #f0f0f0; color: #333; border-style: solid; border-width: 2px; border-color: #ccc; border-radius: 5px; }")
        # self.out_textedit.setFixedSize(200, 700)

        # 创建一个选择栏
        self.combobox = QComboBox()
        self.combobox.setMaximumWidth(100)
        self.combobox.setStyleSheet(
            "QComboBox { background-color: white; color: #333; border-style: solid; border-width: 2px; border-color: #ccc; border-radius: 5px; padding: 3px; }"
            "QComboBox::drop-down { subcontrol-origin: padding; subcontrol-position: top right; width: 20px; border-left-width: 1px; border-left-color: #ccc; border-left-style: solid; background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #f6f7fa, stop:1 #dadbde); }"
            "QComboBox::down-arrow { image: url(down_arrow.png); }")
        # self.combobox.setFixedSize(150, 50)
        # 给下拉列表添加选项
        self.combobox.addItem("模型选择")
        self.combobox.addItems(['EMD-ARIMA', 'ARIMA-RF', 'Kalman-BP', 'LSTM-ARIMA'])

        # 创建一个水平布局器
        self.h_layout = QHBoxLayout()
        self.h_layout.addWidget(self.label1)
        self.h_layout.addWidget(self.dateselect_lineEdit)
        self.h_layout.addWidget(self.label2)

        # 创建一个水平布局器存放时间
        self.time_layout = QHBoxLayout()
        self.time_layout.addWidget(self.year_lineEdit)
        self.time_layout.addWidget(self.label4)
        self.time_layout.addWidget(self.mouth_lineEdit)

        # 创建一个垂直布局器用来选择时间
        self.v_layout1_time = QVBoxLayout()
        self.v_layout1_time.addWidget(self.label3)
        self.v_layout1_time.addLayout(self.time_layout)

        # 创建一个垂直布局器用来存放水位信息
        self.v_layout1_waterlevel = QVBoxLayout()

        self.labelwater1 = QLabel(self)
        self.labelwater1.setMaximumWidth(80)
        self.labelwater1.setText("起始港水位：")
        self.v_layout1_waterlevel.addWidget(self.labelwater1)

        self.water_layout_level1 = QHBoxLayout()
        self.waterlevel1_lineEdit = QLineEdit()
        self.waterlevel1_lineEdit.setMaximumWidth(70)
        self.waterlevel1_lineEdit.setStyleSheet(
            "QLineEdit { background-color: white; color: #333; border-style: solid; border-width: 2px; border-color: #ccc; border-radius: 5px; padding: 3px; }")
        self.water_layout_level1.addWidget(self.waterlevel1_lineEdit)
        self.labelwater3 = QLabel(self)
        self.labelwater3.setMaximumWidth(20)
        self.labelwater3.setText("米")
        self.water_layout_level1.addWidget(self.labelwater3)
        self.v_layout1_waterlevel.addLayout(self.water_layout_level1)

        self.labelwater2 = QLabel(self)
        self.labelwater2.setMaximumWidth(80)
        # self.label1.setFixedSize(50, 20)
        self.labelwater2.setText("目的港水位：")
        self.v_layout1_waterlevel.addWidget(self.labelwater2)

        self.water_layout_level1 = QHBoxLayout()
        self.water2_lineEdit = QLineEdit()
        self.water2_lineEdit.setMaximumWidth(70)
        self.water2_lineEdit.setStyleSheet(
            "QLineEdit { background-color: white; color: #333; border-style: solid; border-width: 2px; border-color: #ccc; border-radius: 5px; padding: 3px; }")
        self.water_layout_level1.addWidget(self.water2_lineEdit)
        self.labelwater4 = QLabel(self)
        self.labelwater4.setMaximumWidth(20)
        # self.label1.setFixedSize(50, 20)
        self.labelwater4.setText("米")
        self.water_layout_level1.addWidget(self.labelwater4)
        self.v_layout1_waterlevel.addLayout(self.water_layout_level1)

        # 创建一个垂直布局器，用来左边控件
        self.v_layout1 = QVBoxLayout()
        self.v_layout1.addWidget(self.but_input)
        self.v_layout1.addStretch(1)
        self.v_layout1.addLayout(self.h_layout)
        # self.v_layout1.addStretch(1)
        self.v_layout1.addWidget(self.but_load)
        self.v_layout1.addStretch(1)
        self.v_layout1.addLayout(self.v_layout1_time)
        self.v_layout1.addStretch(1)
        self.v_layout1.addLayout(self.v_layout1_waterlevel)
        self.v_layout1.addStretch(1)
        self.v_layout1.addWidget(self.combobox)
        self.v_layout1.addStretch(1)
        self.v_layout1.addWidget(self.but_pred)
        self.v_layout1.addStretch(1)
        self.v_layout1.addWidget(self.but_img)
        self.v_layout1.addStretch(1)
        self.v_layout1.addWidget(self.but_nextyear_pre)
        self.v_layout1.addStretch(1)
        self.v_layout1.addWidget(self.but_nextyear_show)

        # 创建一个垂直布局器，用来存放中间控件
        self.v_layout2 = QVBoxLayout()
        self.v_layout2.addWidget(self.image_label)
        self.v_layout2.addWidget(self.textedit)

        # 创建一个垂直布局器，用来存放中间控件
        self.v_layout3 = QVBoxLayout()
        self.v_layout3.addWidget(self.out_textedit)
        self.v_layout3.addWidget(self.but_out)

        # 设置外部布局器
        self.container.addLayout(self.v_layout1)
        self.container.addLayout(self.v_layout2)
        self.container.addLayout(self.v_layout3)
        # self.setCentralWidget(self.container)
        self.setLayout(self.container)

        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        # 连接窗口大小变化的信号与槽
        # self.resizeEvent = self.customResizeEvent

        self.but_input.clicked.connect(self.but_input_clicked)
        self.but_pred.clicked.connect(self.but_pred_clicked)
        self.but_load.clicked.connect(self.but_load_clicked)
        self.but_img.clicked.connect(self.but_clicked_show)
        self.but_out.clicked.connect(self.but_clicked_output)
        self.but_nextyear_pre.clicked.connect(self.but_nextyear_clicked_pre)
        self.but_nextyear_show.clicked.connect(self.but_nextyear_clicked_show)

        self.combobox.currentIndexChanged.connect(self.combobox_clicked)

        self.myThread = MyThread_getdata()  # 调用一个子线程

        self.myThread_ARIMA = MyThread_ARIMA()
        self.myThread_ARIMA_pre = MyThread_ARIMA_pre()

        self.myThread_KalmanBP = MyThread_KalmanBP()
        self.myThread_KalmanBP_pre = MyThread_KalmanBP_pre()

        self.myThread_ARIMABF = MyThread_ARIMABF()
        self.myThread_ARIMABF_pre = MyThread_ARIMABF_pre()

        self.myThread_EMDARIMA = MyThread_EMDARIMA()
        self.myThread_EMDARIMA_pre = MyThread_EMDARIMA_pre()

    # def customResizeEvent(self, event):
    #     # 在窗口大小变化时调用此方法
    #     super(Mywindow, self).resizeEvent(event)
    #
    #     # 根据窗口大小调整部件大小
    #     window_width = self.width()
    #     window_height = self.height()
    #     print(window_height)
    #
    #     # 根据窗口大小调整按钮大小
    #     self.but_input.setMinimumSize(int(window_width / 8), int(window_height / 8))
    #     self.but_pred.setMinimumSize(int(window_width / 8), int(window_height / 8))
    #     self.but_load.setMinimumSize(int(window_width / 8), int(window_height / 8))
    #     self.but_img.setMinimumSize(int(window_width / 8), int(window_height / 8))
    #     # self.but_out.setMinimumSize(window_width / 8, window_height / 8)
    #     self.but_nextyear_pre.setMinimumSize(int(window_width / 8), int(window_height / 8))
    #     self.but_nextyear_show.setMinimumSize(int(window_width / 8), int(window_height / 8))
    #
    #     self.label1.setMinimumSize(int(window_width / 24), int(window_height / 8))
    #     self.label2.setMinimumSize(int(window_width / 24), int(window_height / 8))
    #     self.lineEdit1.setMinimumSize(int(window_width / 24), int(window_height / 8))
    #
    #     self.image_label.setMinimumSize(int((5 * window_width) / 8), int((6 * window_height) / 8))
    #     self.textedit.setMinimumSize(int((5 * window_width) / 8),int((2 * window_height) / 8))
    #
    #     self.out_textedit.setMinimumSize(int((2 * window_width) / 8), int((7 * window_height) / 8))
    #     self.but_out.setMinimumSize(int((2 * window_width) / 8), int((1 * window_height) / 8))

    # 获取当前是哪种模型
    def combobox_clicked(self, i):
        self.selectItems_model = self.combobox.currentText()
        self.textedit.setPlainText("已选择：" + self.selectItems_model + "，训练需要时间，请稍等……")
        if self.selectItems_model == "EMD-ARIMA":
            self.but_img.setStyleSheet(
                "QPushButton { background-color: #c8c8c8; color: white; border-style: outset; border-width: 2px; border-radius: 100px; border-color: #2E8B57; }"
                "QPushButton:hover { background-color: #45a049; }"
                "QPushButton:pressed { background-color: #3e8e41; }")
            self.but_img.setEnabled(False)
        else:
            self.but_img.setStyleSheet(
                "QPushButton { background-color: #4CAF50; color: white; border-style: outset; border-width: 2px; border-radius: 100px; border-color: #2E8B57; }"
                "QPushButton:hover { background-color: #45a049; }"
                "QPushButton:pressed { background-color: #3e8e41; }")
            self.but_img.setEnabled(True)

    def but_input_clicked(self):
        self.filePath, self.filetype = QFileDialog.getOpenFileName(self, "选取文件", "xlsx文件", "*.xlsx")
        self.textedit.setPlainText("已打开文件:" + str(self.filePath))

    def but_load_clicked(self):

        if self.dateselect_lineEdit.text() == "":
            self.line = 2
        else:
            self.line = self.dateselect_lineEdit.text()
        self.textedit.setText("正在加载数据：" + self.filePath.split("/")[-1] + "……")
        print("选择数据{}行".format(self.line))
        # 数据导入
        self.wb = openpyxl.load_workbook(self.filePath.replace("/", "//"))  # 读取文件路径
        # 打开指定的工作簿中的指定工作表：
        self.ws = self.wb["Sheet1"]
        self.ws = self.wb.active  # 打开激活的工作表
        self.ws = list(self.ws.values)  # 转为列表
        self.dates = list(self.ws[0])[5:]
        self.base_information = list(self.ws[int(self.line) - 1])[:5]
        print("基本信息为：",self.base_information)
        self.prices = list(self.ws[int(self.line) - 1])[5:]
        print("价格信息为：",self.prices)
        self.prewaterlevel1 = list(self.ws[int(self.line)])[5:]
        print("起始港水位信息为信息为：",self.prewaterlevel1)
        self.prewaterlevel2 = list(self.ws[int(self.line) + 1])[5:]
        print("目的港水位信息为信息为：",self.prewaterlevel2)



        self.out_textedit.append("基本信息为：")
        self.out_textedit.append(str(self.base_information))

        self.out_textedit.append("油价信息为：")
        for num in range(len(self.dates)):
            self.out_textedit.append(f"{self.dates[num]}：{self.prices[num]}")

        self.myThread.start()  # 执行子线程
        self.myThread.getdata_signal.connect(self.myThread.get_picture)
        self.myThread.getdata_signal.emit(str(self.dates), str(self.prices))

        self.pixmap = QPixmap("output//data_initial.png")  # 按指定路径找到图片
        self.image_label.setPixmap(self.pixmap)  # 在label上显示图片
        self.image_label.setScaledContents(True)  # 让图片自适应label大小

    def but_clicked_show(self):
        if os.path.exists("output//out_pre.png") and os.path.exists("out.txt"):
            self.pixmap = QPixmap("output//out_pre.png")  # 按指定路径找到图片
            self.image_label.setPixmap(self.pixmap)  # 在label上显示图片
            self.image_label.setScaledContents(True)  # 让图片自适应label大小

            self.out_textedit.setPlainText("对原始数据拟合结果为：")
            with open("out.txt", 'r+') as f:
                self.content = f.read()

            self.out_textedit.append(self.content)
            self.textedit.setPlainText("输出模型训练结果！")

        else:
            self.textedit.setPlainText("暂无训练结果，请加载数据并选择模型进行训练并预测！")

    def but_nextyear_clicked_show(self):
        if os.path.exists("output//nextyear_pre.png") and os.path.exists("out_pre.txt"):
            self.pixmap_1 = QPixmap("output//nextyear_pre.png")  # 按指定路径找到图片
            self.image_label.setPixmap(self.pixmap_1)  # 在label上显示图片
            self.image_label.setScaledContents(True)  # 让图片自适应label大小

            self.out_textedit.setPlainText("预测结果为：")
            with open("out_pre.txt", 'r+') as f:
                self.content2 = f.read()

            self.out_textedit.append(self.content2)
            self.textedit.setPlainText("输出模型预测结果！")

        else:
            self.textedit.setPlainText("暂无预测结果，请加载数据并选择模型进行预测！")


    ##开始训练模型
    def but_pred_clicked(self):
        """
        得到时间值
        """
        self.year = self.year_lineEdit.text()
        self.mouth = self.mouth_lineEdit.text()
        self.date_time = self.year + "-" + self.mouth
        """
        得到水位信息
        """
        self.waterlevel1 = self.waterlevel1_lineEdit.text()
        self.waterlevel2 = self.water2_lineEdit.text()

        if self.selectItems_model == "EMD-ARIMA":
            #'EMD-ARIMA', 'ARIMA-RF', 'Kalman-BP', 'LSTM-ARIMA'
            self.textedit.setPlainText("这是模型一：EMD-ARIMA")
            self.myThread_EMDARIMA.start()  # 执行子线程
            self.myThread_EMDARIMA.EMDARIMA_signal.connect(self.myThread_EMDARIMA.get_model)
            self.myThread_EMDARIMA.EMDARIMA_signal.emit(str(self.dates), str(self.prices), self.date_time,  self.waterlevel1, self.waterlevel2)
            self.textedit.setPlainText("模型训练完毕！")

        elif self.selectItems_model == "ARIMA-RF":
            self.textedit.setPlainText("这是模型二：ARIMA-RF")
            # self.textedit.setPlainText("这是模型三！")
            self.myThread_ARIMABF.start()  # 执行子线程
            self.myThread_ARIMABF.ARIMABF_signal.connect(self.myThread_ARIMABF.get_model)
            self.myThread_ARIMABF.ARIMABF_signal.emit(str(self.dates), str(self.prices), self.date_time,  self.waterlevel1, self.waterlevel2)
            self.textedit.setPlainText("模型训练完毕！")

        elif self.selectItems_model == "Kalman-BP":
            self.textedit.setPlainText("这是模型三：Kalman-BP")
            self.myThread_KalmanBP.start()  # 执行子线程
            self.myThread_KalmanBP.KalmanBP_signal.connect(self.myThread_KalmanBP.get_model)
            self.out = str(self.dates) + "*" + str(
                self.prices) + "*" + self.date_time + "*" + self.waterlevel1 + "*" + self.waterlevel2 + "*" + str(
                self.prewaterlevel1) + "*" + str(self.prewaterlevel2)
            self.myThread_KalmanBP.KalmanBP_signal.emit(self.out)
            self.textedit.setPlainText("模型训练完毕！")

        elif self.selectItems_model == "LSTM-ARIMA":
            self.textedit.setPlainText("这是模型四：LSTM-ARIMA")
            self.myThread_ARIMA.start()  # 执行子线程
            self.myThread_ARIMA.ARIMA_signal.connect(self.myThread_ARIMA.get_model)
            self.myThread_ARIMA.ARIMA_signal.emit(str(self.dates), str(self.prices), self.date_time,  self.waterlevel1, self.waterlevel2)
            self.textedit.setPlainText("模型训练完毕！")

        else:
            self.textedit.setPlainText("请选择预测模型！")

    ##开始预测
    def but_nextyear_clicked_pre(self):
        """
        得到时间值
        """
        self.year = self.year_lineEdit.text()
        self.mouth = self.mouth_lineEdit.text()
        self.day = self.day_lineEdit.text()
        self.date_time = self.year + "-" + self.mouth + "-" + self.day
        print(self.date_time)

        if self.selectItems_model == "EMD-ARIMA":
            self.textedit.setPlainText("这是模型一：EMD-ARIMA")
            self.myThread_EMDARIMA_pre.start()  # 执行子线程
            self.myThread_EMDARIMA_pre.EMDARIMA_pre_signal.connect(self.myThread_EMDARIMA_pre.model_pre)
            self.myThread_EMDARIMA_pre.EMDARIMA_pre_signal.emit(str(self.dates), str(self.prices))
            self.textedit.setPlainText("模型预测完毕！")

        elif self.selectItems_model == "ARIMA-RF":
            self.textedit.setPlainText("这是模型二：ARIMA-RF")
            self.myThread_ARIMABF_pre.start()  # 执行子线程
            self.myThread_ARIMABF_pre.ARIMABF_pre_signal.connect(self.myThread_ARIMABF_pre.model_pre)
            self.myThread_ARIMABF_pre.ARIMABF_pre_signal.emit(str(self.dates), str(self.prices))
            self.textedit.setPlainText("模型预测完毕！")


        elif self.selectItems_model == "Kalman-BP":
            self.textedit.setPlainText("这是模型三：Kalman-BP")
            self.myThread_KalmanBP_pre.start()  # 执行子线程
            self.myThread_KalmanBP_pre.KalmanBP_pre_signal.connect(self.myThread_KalmanBP_pre.model_pre)
            self.myThread_KalmanBP_pre.KalmanBP_pre_signal.emit(str(self.dates), str(self.prices))
            self.textedit.setPlainText("模型预测完毕！")

        elif self.selectItems_model == "LSTM-ARIMA":
            self.textedit.setPlainText("这是模型四：LSTM-ARIMA")
            self.myThread_ARIMA_pre.start()  # 执行子线程
            self.myThread_ARIMA_pre.ARIMA_pre_signal.connect(self.myThread_ARIMA_pre.model_pre)
            self.myThread_ARIMA_pre.ARIMA_pre_signal.emit(str(self.dates), str(self.prices))
            self.textedit.setPlainText("模型预测完毕！")

        else:
            self.textedit.setPlainText("请选择预测模型！")

    def but_clicked_output(self):
        file_path = "output.zip"
        save_file_path, _ = QFileDialog.getSaveFileName(self, "保存文件", "", "Excel Files (*.zip)")
        if save_file_path:
            with zipfile.ZipFile(save_file_path, 'w') as zipf:
                zipf.write(file_path)
            print("文件已成功导出到:", save_file_path)

        self.textedit.setPlainText("文件成功导出！")


if __name__ == '__main__':
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps)
    app = QApplication(sys.argv)
    w = Mywindow()

    w.show()

    app.exec()